from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl

# updating invoices enmass to flag the 'paid' flag to checked (yes) for
# a list of invoices in a spreadsheet that are not already flagged

StartTime = time.time() 

wb = openpyxl.load_workbook('Invoice_Update_PaidFlag_Input.xlsx')
shx = wb.get_sheet_by_name('Login')
Login = shx.cell(row=2, column=1).value
Password = shx.cell(row=2, column=2).value
clientURL = shx.cell(row=2, column = 3).value

sh = wb.get_sheet_by_name('Sheet1')
MaxRow = sh.max_row

driver = webdriver.Chrome()
driver.get(clientURL)
driver.implicitly_wait(10)

# >>>>>>>>>>>>>> LOGIN >>>>>>>>>>>>>>>
loginElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_login'))
userPasswordElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_password'))
buttonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name('button'))

loginElement.send_keys(Login)
userPasswordElement.send_keys(Password)
buttonElement.click()
# <<<<<<<<<<<<<< LOGIN <<<<<<<<<<<<<<<<

# >>>>>>>>>>> navigate to invoice >>>>>>>>>
invoiceTabElement = WebDriverWait(driver, 10).until(
    lambda driver: driver.find_element_by_xpath("//a[@href='/invoices']"))
invoiceTabElement.click()
# <<<<<<<<<<< navigate to invoice <<<<<<<<<<

# >>>>>> advanced search >>>>>>>>>
advancedSearchElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('si_invoice_header'))
advancedSearchElement.click()
advSearchElements = WebDriverWait(driver, 10).until(
    lambda driver: driver.find_elements_by_xpath("//div[@id='invoice_header_adv_cond_w']"
                                                 "/div[@class='condition_row']"))
# print(len(advSearchElements))
firstSearchElement = advSearchElements[0]
firstPicklistInvoiceElement = firstSearchElement.find_element_by_xpath(
    "//option[@value='custom_field_16']")  # clicking 'Legacy PO' filter
firstPicklistInvoiceElement.click()
time.sleep(2)
firstPickListConditionClauseElement = firstSearchElement.find_element_by_xpath("//option[@value='not_blank']")
firstPickListConditionClauseElement.click()
time.sleep(2)
picklistAddElement = firstSearchElement.find_element_by_xpath("//a[@class='invoice_header_wait condition_add_link']")
picklistAddElement.click()
WebDriverWait(driver, 10).until(
    lambda driver: len(driver.find_elements_by_xpath("//div[@id='invoice_header_adv_cond_w']"
                                                     "/div[@class='condition_row']")) == 2)
advSearchElements2 = WebDriverWait(driver, 10).until(
    lambda driver: driver.find_elements_by_xpath("//div[@id='invoice_header_adv_cond_w']"
                                                 "/div[@class='condition_row']"))
# print(len(advSearchElements2))
secondSearchElement = advSearchElements2[1]
secondPicklistInvoiceElement = secondSearchElement.find_element_by_xpath(".//option[@value='invoice_number']")
secondPicklistInvoiceElement.click()
inputElement = secondSearchElement.find_element_by_xpath(".//input[@type='text']")
inputElement.clear()
inputElement.send_keys('PO-Migration' + '\n')
# there is a need to search and click into an existing invoice for rest of script to work
# this is because the adv search filters are saved ONLY when you click into an item that
# was produced initially from the search
time.sleep(2)
firstInvoiceElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(
    "//tr[contains(@id,'invoice_header_row')]//span[@class='dt_open_link']"))
firstInvoiceElement.click()
time.sleep(2)
# <<<<<<< advanced search <<<<<<<<<<


# >>>>>>>>>>>>>>>>>>>>> Click into invoice and update 'Paid' flag to YES >>>>>>>>>>>>>>>>>>>>
for x in range(2,MaxRow+1):
    invoiceSearchInput = sh.cell(row=x, column=1).value
    invoiceURL = shx.cell(row=2, column=4).value
    invoiceDateInput = sh.cell(row=x,column=2).value
    print('Invoice: ' + str(invoiceSearchInput))
    # driver.get(invoiceURL) # navigate to invoice page // commenting out for now due to clicking on invoice tab instead (below 2 lines)
    invoiceTabElement = WebDriverWait(driver, 10).until(
        lambda driver: driver.find_element_by_xpath(
            "//nav[@class='primary']//a[@class='on'][@href='/invoices']"))
    invoiceTabElement.click()

    advSearchElements2 = WebDriverWait(driver, 10).until(
        lambda driver: driver.find_elements_by_xpath("//div[@id='invoice_header_adv_cond_w']"
                                                     "/div[@class='condition_row']"))
    print(len(advSearchElements2))
    secondSearchElement = advSearchElements2[1]
    inputElement = secondSearchElement.find_element_by_xpath(".//input[@type='text']")
    inputElement.clear()
    inputElement.send_keys(str(invoiceSearchInput) + '\n')

    time.sleep(2) # wait for invoice elements to refresh properly
    firstInvoiceElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(
        "//tr[contains(@id,'invoice_header_row')]//span[@class='dt_open_link']"))
    firstInvoiceElements = WebDriverWait(driver, 10).until(lambda driver: driver.find_elements_by_xpath(
        "//tr[contains(@id,'invoice_header_row')]//span[@class='dt_open_link']"))
    print('Number of Invoices from search: ' + str(len(firstInvoiceElements)))
    time.sleep(1)
    WebDriverWait(driver, 2).until(lambda driver: invoiceSearchInput in driver.find_element_by_xpath(
        "//tr[contains(@id,'invoice_header_row')]//span[@class='dt_open_link']").text) # wait 2s until the invoice # is inside the first search

    if len(firstInvoiceElements) == 1: # if only 1 invoice is found form search then click into invoice
        firstInvoiceElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(
            "//tr[contains(@id,'invoice_header_row')]//span[@class='dt_open_link']"))
        firstInvoiceElement.click()

        # >>>>>>>>>>>> Assert the invoice clicked is accurate >>>>>>>>>>
        invoicePageHeaderElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('pageHeader'))
        invoicePageHeader = invoicePageHeaderElement.text
        assert invoiceSearchInput in invoicePageHeader
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>> selecting 'paid' >>>>>>>>>>>>
        invoicePaidElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('invoice_paid'))
        invoiceDateElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('invoice_local_payment_date'))
        sh.cell(row=x, column=4).value = invoicePaidElement.is_selected()
        print(invoicePaidElement.is_selected())
        if invoicePaidElement.is_selected() is False:  # Update this to TRUE if wanting to change to Paid flag as no
            invoicePaidElement.click()  # checkbox for 'paid'
            invoiceDateElement.clear()
            invoiceDateElement.send_keys(str(invoiceDateInput) + '\n')
            savePaymentDetailsElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//a[@class=' rollover button ']//img[@class='icon icon_button sprite-disk']"))
            savePaymentDetailsElement.click()
            WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('pageHeader').text == 'Invoices')
            # wait for invoice page to load = previous invoice was saved
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

            sh.cell(row=x, column=3).value = 'Paid checkbox configured'
            wb.save('Invoice_Update_PaidFlag_Output.xlsx')
            print(time.time() - StartTime)
        else:
            sh.cell(row=x, column=3).value = 'Paid checkbox was already configured'
            wb.save('Invoice_Update_PaidFlag_Output.xlsx')
    else:
        sh.cell(row=x, column=3).value = 'More than 1 invoice returned from search'
        wb.save('Invoice_Update_PaidFlag_Output.xlsx')
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

wb.save('Invoice_Update_PaidFlag_Output.xlsx')
driver.quit()
print(time.time() - StartTime)













