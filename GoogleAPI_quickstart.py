from __future__ import print_function
import httplib2
import os
from apiclient import errors
from apiclient import discovery
import oauth2client
from oauth2client import client
from oauth2client import tools
import base64
import email
import re
import openpyxl

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# This script is used to pull the password reset emails from a specified GMAIL account and
# write them to a spreadsheet.
# Step 1: log into coupa and in Setup>Users import a list of existing or new users with the flag
#         of 'reset password' as "Yes"
# Step 2: after import is complete, log into the gmail account used to reset the users (ie. all
#         users must have the same Gmail account such as Coupa+xxx@gmail.com). Manually move
#         all of the password emails sent by Coupa into a specified folder. In the example, Label_5
#         is the ID of the folder used.
# Step 3: Run the below script and an output file will be created with all the links contained
# Step 4: Run the 'User_ChangePassword.py' script to loop through all the links and set the
#         password to a defaulted password


wb = openpyxl.load_workbook('GoogleAPI_quickstart_Output.xlsx')
sh = wb.get_sheet_by_name('Sheet1')

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/gmail-python-quickstart.json

SCOPES = 'https://www.googleapis.com/auth/gmail.readonly'
# SCOPES = 'https://www.googleapis.com/auth/gmail.modify'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Gmail API Python Quickstart'


def get_credentials():
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'gmail-python-quickstart.json')

    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def main():
    """Shows basic usage of the Gmail API.

    Creates a Gmail API service object and outputs a list of label names
    of the user's Gmail account.
    """
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('gmail', 'v1', http=http)
    results = service.users().labels().list(userId='me').execute()
    labels = results.get('labels', [])
    max_Results = ''
    labels_ID = 'Label_5' # specify the label ID after pulling it via the label prints to console below
    query = ''

    if not labels:
        print('No labels found.')
    else:
        print('Labels:')
        for label in labels:
            print(label['id'] + ' ' + label['name'])

    response = service.users().messages().list(userId='me',
                                               labelIds = labels_ID
                                               ).execute()
    messages = []
    if 'messages' in response:
        messages.extend(response['messages'])

    while 'nextPageToken' in response:
        page_token = response['nextPageToken']
        response = service.users().messages().list(userId='me',
                                                   pageToken=page_token,
                                                   labelIds = labels_ID
                                                   ).execute()
        messages.extend(response['messages'])

    rowNum = 1 # starting row of the spreadsheet
    for x in messages: # the piece of code to pull and parse the messages
        try:
            y = service.users().messages().get(userId='me', id=x['id'],format='raw').execute()
            z = base64.urlsafe_b64decode(y['raw'].encode('ASCII'))
            zz = str(z)
            # print(z) # comment this in to get entire string. Below is to parse out the required link
            linkBeginningTextLocation = re.search('https',zz).start() # starting point of where required link sits in string
            linkEndingTextLocation = re.search('Your access credentials',zz).start() # ending point of where link sits in string
            # print(z[linkBeginningTextLocation:linkEndingTextLocation])
            sh.cell(row=rowNum, column=1).value = zz[linkBeginningTextLocation:linkEndingTextLocation]
            accountBeginningTextLocation = re.search('Username: ',zz).start() # starting point of where required link sits in string
            accountEndingTextLocation = re.search('Web Address:',zz).start() # ending point of where link sits in string
            # print(z[linkBeginningTextLocation:linkEndingTextLocation])
            sh.cell(row=rowNum, column=2).value = zz[accountBeginningTextLocation:accountEndingTextLocation]
            rowNum += 1
        except:
            rowNum += 1
            pass


    wb.save('GoogleAPI_quickstart_Output.xlsx')

if __name__ == '__main__':
    main()
