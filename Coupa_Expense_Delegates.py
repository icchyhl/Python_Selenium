from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
import time
import openpyxl

# this script loops through the IDs in Coupa Expense Delegates excel file and does something
StartTime = time.time()

driver = webdriver.Chrome()
driver.get('https://deloitte-ca2.coupacloud.com/session')
# driver.implicitly_wait(1)

wbx = openpyxl.load_workbook('Login.xlsx')
shx = wbx.get_sheet_by_name('Sheet1')
Login = shx.cell(row=1, column=1).value
Password = shx.cell(row=1, column=2).value

loginElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_login'))
userPasswordElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_password'))
buttonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name('button'))

loginElement.send_keys(Login)
userPasswordElement.send_keys(Password)
buttonElement.click()

wb = openpyxl.load_workbook('Coupa_Expense_Delegates_Input.xlsx')
sh = wb.get_sheet_by_name('Sheet1')
MaxRow = sh.max_row

print("Max Row = %s" % MaxRow)

for x in range(2,MaxRow+1):
    AccountID = sh.cell(row=x, column=1).value
    print(AccountID)
    driver.get('https://deloitte-ca2.coupacloud.com/user/edit_user/%s' % AccountID)
    SampleText = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('pageHeader'))
    print(SampleText)

    FullName = sh.cell(row=x, column=2).value
    fullNameElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_can_delegate_expenses_to_ids_name'))
    fullNameElement.send_keys(FullName)

    print(time.time() - StartTime)

    for y in range(1,10):
        try:
            delegateIDElement = driver.find_element_by_id('ui-id-%s' % y)
            print(delegateIDElement.text)
            if sh.cell(row=x,column=3).value in delegateIDElement.text:
                delegateIDElement.click()
                sh.cell(row=x, column=4).value = 'Delegate Assigned'
                submitButtonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//button[@class='button blue'][@type='submit']"))
                submitButtonElement.click()
        except:
            pass

    wb.save('Coupa_Expense_Delegate_Output.xlsx')

driver.close()
wb.save('Coupa_Expense_Delegate_Output.xlsx')
print(time.time() - StartTime)