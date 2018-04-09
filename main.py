from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import unittest
from unittest import runner
import time

# Test Automation scripts for Coupa R15

class TestResult(runner.TextTestResult):
    """
    Used to show the different test results and mark it into Excel
    """
    def addError(self, test, err):
        test.markCell('Error')
        super(TestResult, self).addError(test, err)

    def addFailure(self, test, err):
        test.markCell('Failure')
        super(TestResult, self).addFailure(test, err)

    def addSuccess(self, test):
        test.markCell('Success')
        super(TestResult, self).addSuccess(test)

class test_0000_Config(unittest.TestCase):
    """
    Use this to configure anything required at the very beginning before any test scenarios are run
    """
    def test_spreadsheet_Config(self):
        self.wb = openpyxl.load_workbook('Main_Input.xlsx')
        self.wb.save('Main_Output.xlsx')

    def markCell(self, value):
        pass # script needs to call markCell for config to work

class test_0020_SearchCatalogs(unittest.TestCase):
    """
    Test Scenario for (RC21): "Search existing catalogs "
    """
    @classmethod
    def setUpClass(cls):
        # this will set up the initial values for this class, which the
        # test runner will do.
        cls.wb = openpyxl.load_workbook('Main_Output.xlsx')
        if cls.wb.get_sheet_by_name('Requisition Creation').cell(row=4,column=10).value == "Yes":
            cls.skipTest(cls,test_0010_Login) # Skip entire test scenario if
        cls.sh_Setup = cls.wb.get_sheet_by_name('Setup')
        cls.ClientURL = cls.sh_Setup.cell(row=1, column=2).value
        cls.driver = webdriver.Chrome()
        cls.driver.get(cls.ClientURL)

    def tearDown(self):
        # This will save the progress if test scenario is aborted before full execution
        self.wb.save('Main_Output.xlsx')

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()
        cls.wb.save('Main_Output.xlsx')

    def sheetLocation(self, sheet, row, col):
        # initialize the initial value to success first
        self._sh, self._row, self._col = sheet, row, col
        self.sh = self.wb.get_sheet_by_name(self._sh)

    def markCell(self, value):
        self.sh.cell(row=self._row, column=self._col).value = value

    # ============= Beginning of the steps from the test scenario start from below ===========

    def test_0100_Login(self):
        """
        step "RC21.1": "Login to Coupa, if you are already logged in,
        Please go to home page of Coupa by clicking home icon right beneath the Company logo"
        """
        self.sheetLocation('Requisition Creation',5,8)

        driver = self.driver
        coupaLogin = self.sh_Setup.cell(row=4, column=1).value
        coupaPassword = self.sh_Setup.cell(row=4, column=2).value
        loginFieldID = 'user_login'
        passwordFieldID = 'user_password'
        loginButtonClassName = 'button'
        coupaLogoID = 'global-logo'
        homeURL = 'user/home'

        loginFieldElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(loginFieldID))
        passwordFieldElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(passwordFieldID))
        loginButtonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name(loginButtonClassName))

        loginFieldElement.clear()
        loginFieldElement.send_keys(coupaLogin)
        passwordFieldElement.clear()
        passwordFieldElement.send_keys(coupaPassword)
        loginButtonElement.click()
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(coupaLogoID))
        assert homeURL in driver.current_url # ensure URL is on the home page

    def test_0200_OpenCatalog(self):
        """
        step "RC21.2": On the home page, right under blue ribbon and besides "Webform" hover over "Catalogs"
        """
        self.sheetLocation('Requisition Creation', 6, 8)

        driver = self.driver
        catalogDropdownXpath = "//div[@class='purchasing_menu_container catalogs']/a[@class='purchasing_menu_link']"
        catalogItemsXpath = "//div[@class='purchasing_menu_container catalogs']/ul[@id='catalogs_menu']//a[@*]"

        catalogListElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(catalogDropdownXpath))
        catalogListElement.click()
        catalogItemsElements = WebDriverWait(driver, 10).until(lambda driver: driver.find_elements_by_xpath(catalogItemsXpath))
        assert len(catalogItemsElements) > 0 # ensure there are 1 or more categories to pass the test

    def test_0300_SelectCatalog(self):
        """
        step "RC21.3": Click on a catalog named in "Test Input1"
        """
        self.sheetLocation('Requisition Creation', 7, 8)
        wb = self.wb
        inputSheet = wb.get_sheet_by_name('Requisition Creation')
        testInput1 = inputSheet.cell(row=7,column=11).value

        driver = self.driver
        catalogTitleXpath = "//div[@class='purchasing_menu_container catalogs']/ul[@id='catalogs_menu']//a[@title='" + testInput1 + "']"
        pageHeaderID = 'pageHeader'

        catalogTitleElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(catalogTitleXpath))
        catalogTitleElement.click()
        pageHeaderElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(pageHeaderID))
        assert testInput1 in pageHeaderElement.text # the page header should have the catagory name indicated in the spreadsheet input

    def test_0400_NextPage(self):
        """
        step "RC21.4": Scroll down and on the lower right corner of the page, click on "Next"
        """
        self.sheetLocation('Requisition Creation', 8, 8)

        driver = self.driver
        nextButtonClass = "next_page"
        currentPageClass = 'current'

        nextButtonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name(nextButtonClass))
        nextButtonElement.click()
        currentPageElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name(currentPageClass))
        assert currentPageElement.text == '2' # upon clicking next page, the current page should indicate "2"

    def test_0500_SelectItem(self):
        """
        step "RC21.5": Click on the first item to see the details
        """
        self.sheetLocation('Requisition Creation', 9, 8)

        driver = self.driver
        firstItemXpath = "//div[@class='item-title'][1]"
        itemDescriptionXpath = "//div[@id='item_content']//div[@class='item_description']/div[@class='item_info']"

        firstItemElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(firstItemXpath))
        firstItemElement.click()
        itemDescriptionElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(itemDescriptionXpath))
        assert len(itemDescriptionElement.text) > 0 # ensure there is content in the item description area

    def test_0600_AddtoCart(self):
        """
        step "RC21.6": Click "add to Cart" to add this item in your cart
        """
        self.sheetLocation('Requisition Creation', 10, 8)

        driver = self.driver
        addtoCartXpath = "//button[@class='button green'][@type='submit']"
        selectCartID = 'cart'
        cartHeaderID = 'pageHeader'
        itemNameXpath = "//span[@class='info']"

        addtoCartElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath(addtoCartXpath))
        addtoCartElement.click()
        selectCartElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(selectCartID))
        selectCartElement.click()
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(cartHeaderID))
        itemNameElements =  WebDriverWait(driver, 10).until(lambda driver: driver.find_elements_by_xpath(itemNameXpath))
        assert len(itemNameElements) > 0 # ensure cart has 1 or more items

class test_0010_Login(unittest.TestCase):
    """
    Test Scenario for (RC21): "Search existing catalogs "
    """

    @classmethod
    def setUpClass(cls):
        # this will set up the initial values for this class, which the
        # test runner will do.
        cls.wb = openpyxl.load_workbook('Main_Output.xlsx')
        if cls.wb.get_sheet_by_name('User Login').cell(row=4,column=10).value == "Yes":
            cls.skipTest(cls,test_0010_Login)
        cls.sh_Setup = cls.wb.get_sheet_by_name('Setup')
        cls.ClientURL = cls.sh_Setup.cell(row=1, column=2).value
        cls.driver = webdriver.Chrome()
        cls.driver.get(cls.ClientURL)

    def tearDown(self):
        # This will save the progress if test scenario is aborted before full execution
        self.wb.save('Main_Output.xlsx')

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()
        cls.wb.save('Main_Output.xlsx')

    def sheetLocation(self, sheet, row, col):
        # initialize the initial value to success first
        self._sh, self._row, self._col = sheet, row, col
        self.sh = self.wb.get_sheet_by_name(self._sh)

    def markCell(self, value):
        self.sh.cell(row=self._row, column=self._col).value = value

    # ============= Beginning of the steps from the test scenario start from below ===========

    def test_0100_Login(self):
        """
        step "RC21.1": "Login to Coupa, if you are already logged in,
        Please go to home page of Coupa by clicking home icon right beneath the Company logo"
        """
        self.sheetLocation('User Login',5,8)

        driver = self.driver
        coupaLogin = self.sh_Setup.cell(row=4, column=1).value
        coupaPassword = self.sh_Setup.cell(row=4, column=2).value
        loginFieldID = 'user_login'
        passwordFieldID = 'user_password'
        loginButtonClassName = 'button'
        coupaLogoID = 'global-logo'
        homeURL = 'user/home'

        loginFieldElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(loginFieldID))
        passwordFieldElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(passwordFieldID))
        loginButtonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name(loginButtonClassName))

        loginFieldElement.clear()
        loginFieldElement.send_keys(coupaLogin)
        passwordFieldElement.clear()
        passwordFieldElement.send_keys(coupaPassword)
        loginButtonElement.click()
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id(coupaLogoID))
        assert homeURL in driver.current_url # ensure URL is on the home page

if __name__ == "__main__":
    # unittest.main(defaultTest='test_0010_Login', testRunner=runner.TextTestRunner(resultclass=TestResult))
    # unittest.main(defaultTest='test_0020_SearchCatalogs', testRunner=runner.TextTestRunner(resultclass=TestResult))
    unittest.main(testRunner=runner.TextTestRunner(resultclass=TestResult))
    # unittest.main(defaultTest='test_0010_Login')