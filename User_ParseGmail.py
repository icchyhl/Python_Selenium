from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl

# updating invoices enmass to flag the 'paid' flag to checked (yes) for
# a list of invoices in a spreadsheet that are not already flagged

def test():
    StartTime = time.time()

    wb = openpyxl.load_workbook('User_ParseGmail_Output.xlsx')
    shx = wb.get_sheet_by_name('Login')
    Email = shx.cell(row=2, column=1).value
    Password = shx.cell(row=2, column=2).value

    driver = webdriver.Chrome()
    driver.get('https://accounts.google.com/login')
    driver.implicitly_wait(10)

    # >>>>>>>>>>>>>> LOGIN >>>>>>>>>>>>>>>
    emailElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('Email'))
    emailElement.send_keys(Email + '\n')
    userPasswordElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('Passwd'))
    userPasswordElement.send_keys(Password + '\n')
    driver.get('https://mail.google.com/mail')
    # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

if __name__ == '__main__':
    test()
