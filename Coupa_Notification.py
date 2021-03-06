from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl

# this example loops through an excel file and logs in with the information from
# Columns A and B and then writes a value into Column C saving it in a new excel file
# it demonstrates logging in and out of Coupa
# it also sets notifications for the user either ON or OFF depending on the true/false in
# the "select the notifications online & email" section of the code below

StartTime = time.time()

driver=webdriver.Chrome()

wb = openpyxl.load_workbook('Coupa_Notifications_Input.xlsx')
sh = wb.get_sheet_by_name('Sheet1')
MaxRow = sh.max_row

for x in range(2,MaxRow+1):
    print(sh.cell(row=x, column=1).value)
    Login = sh.cell(row=x, column=1).value
    Password = sh.cell(row=x, column=2).value
    driver.get('https://deloitte-ca2.coupacloud.com/session')

    try:
        loginElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_login'))
        passwordElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_password'))
        buttonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name('button'))

        loginElement.send_keys(Login)
        passwordElement.send_keys(Password)
        buttonElement.click()

        driver.get('https://deloitte-ca2.coupacloud.com/inbox/preferences')
    except:
        sh.cell(row=x, column=5).value = "Could not log in"

    try:
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//div[@class='section'][1]"
                                                                        "/div[@class='inline_form_element']"
                                                                        "/input[1]"))
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//div[@class='section'][1]"
                                                                        "/div[@class='inline_form_element']"
                                                                        "/input[2]"))

        Requisition_Notification_Online = driver.find_element_by_xpath("//div[@class='section'][1]"
                                                                        "/div[@class='inline_form_element']"
                                                                        "/input[1]").is_selected()
        Requisition_Notification_Email = driver.find_element_by_xpath("//div[@class='section'][1]"
                                                                       "/div[@class='inline_form_element']"
                                                                       "/input[2]").is_selected()

        print(Requisition_Notification_Online)
        print(Requisition_Notification_Email)

        # ===== select the notifications online & email =====
        if Requisition_Notification_Online is True:
            notification1 = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//div[@class='section'][1]/div[@class='inline_form_element']/input[1]"))
            notification1.click()
        if Requisition_Notification_Email is True:
            notification2 = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//div[@class='section'][1]/div[@class='inline_form_element']/input[2]"))
            notification2.click()
        # ====================================================

        Requisition_Notification_Online_Element = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//div[@class='section'][1]"
                                                                                                                            "/div[@class='inline_form_element']"
                                                                                                                            "/input[1]"))
        Requisition_Notification_Online_Element.is_selected()


        Requisition_Notification_Email_Element = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//div[@class='section'][1]"
                                                                                                                           "/div[@class='inline_form_element']"
                                                                                                                           "/input[2]"))
        Requisition_Notification_Email_Element.is_selected()

        print(Requisition_Notification_Online_Element.is_selected())
        print(Requisition_Notification_Email_Element.is_selected())

        blueButtonElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_xpath("//button[@class='button blue'][@type='submit']"))
        blueButtonElement.click()
        sh.cell(row=x, column=3).value = "Configured"

    except:
        sh.cell(row=x, column=3).value = "NOT configured"
        pass

    try:
        # ==== LOG OUT =========
        myAccountElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('my_account'))
        logoutHover = ActionChains(driver).move_to_element(myAccountElement)
        logoutHover.perform()
        time.sleep(1)
        logoutElement =  WebDriverWait(driver, 10).until(
            lambda driver: driver.find_element_by_xpath("//a[@title='Sign Out'][@href='/sessions/destroy']"))
        logoutElement.click()
        # ======================
    except:
        sh.cell(row=x, column=4).value = "Could not log out"
        pass

    print(time.time() - StartTime)
    wb.save('Notifications_OutputX.xlsx')

driver.quit()
wb.save('Notifications_OutputX.xlsx')
print(time.time() - StartTime)