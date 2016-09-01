from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl

# loops through an output file generated from GoogleAPI_quickstart.py which contains several
# links to reset an email address and sets all the passwords to a default (ie. Password)

StartTime = time.time()
driver = webdriver.Chrome()
wb = openpyxl.load_workbook('GoogleAPI_quickstart_Output.xlsx')
sh = wb.get_sheet_by_name('Sheet1')
MaxRow = sh.max_row
setPassword = 'Password123'

for x in range(1,MaxRow+1):
    print(time.time() - StartTime)
    try:
        passwordResetLink = sh.cell(row=x, column=1).value
        driver.get(passwordResetLink)
        passwordElememt = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_password'))
        passwordConfirmationElememt = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_password_confirmation'))
        submitElement = WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_class_name('button '))
        passwordElememt.clear()
        passwordConfirmationElememt.clear()
        passwordElememt.send_keys(setPassword)
        passwordConfirmationElememt.send_keys(setPassword)
        submitElement.click()
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_login')) # wait until login/password page loads
        WebDriverWait(driver, 10).until(lambda driver: driver.find_element_by_id('user_password'))
        sh.cell(row=x, column=3).value = "Password Set"
    except:
        sh.cell(row=x, column=3).value = "Error: Password not set"
        pass

wb.save('GoogleAPI_quickstart_Output.xlsx')
driver.quit()
print(time.time() - StartTime)